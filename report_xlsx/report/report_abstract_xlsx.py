# Copyright 2015 ACSONE SA/NV (<http://acsone.eu>)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).
import base64
import logging
import re
from io import BytesIO
from urllib.parse import quote

from dateutil.relativedelta import relativedelta
from openpyxl.reader.excel import load_workbook

from odoo import models, fields

_logger = logging.getLogger(__name__)

try:
    import xlsxwriter

    class PatchedXlsxWorkbook(xlsxwriter.Workbook):
        def _check_sheetname(self, sheetname, is_chartsheet=False):
            """We want to avoid duplicated sheet names exceptions the same following
            the same philosophy that Odoo implements overriding the main library
            to avoid the 31 characters limit triming the strings before sending them
            to the library.

            In some cases, there's not much control over this as the reports send
            automated data and the potential exception is hidden underneath making it
            hard to debug the original issue. Even so, different names can become the
            same one as their strings are trimmed to those 31 character limit.

            This way, once we come across with a duplicated, we set that final 3
            characters with a sequence that we evaluate on the fly. So for instance:

            - 'Sheet name' will be 'Sheet name~01'
            - The next 'Sheet name' will try to rename to 'Sheet name~01' as well and
              then that will give us 'Sheet name~02'.
            - And the next 'Sheet name' will try to rename to 'Sheet name~01' and then
              to 'Sheet name~02' and finally it will be able to 'Sheet name~03'.
            - An so on as many times as duplicated sheet names come to the workbook up
              to 100 for each sheet name. We set such limit as we don't want to truncate
              the strings too much and keeping in mind that this issue don't usually
              ocurrs.
            """
            try:
                return super()._check_sheetname(sheetname, is_chartsheet=is_chartsheet)
            except xlsxwriter.exceptions.DuplicateWorksheetName:
                pattern = re.compile(r"~[0-9]{2}$")
                duplicated_secuence = (
                    re.search(pattern, sheetname) and int(sheetname[-2:]) or 0
                )
                # Only up to 100 duplicates
                deduplicated_secuence = "~{:02d}".format(duplicated_secuence + 1)
                if duplicated_secuence > 99:
                    raise xlsxwriter.exceptions.DuplicateWorksheetName
                if duplicated_secuence:
                    sheetname = re.sub(pattern, deduplicated_secuence, sheetname)
                elif len(sheetname) <= 28:
                    sheetname += deduplicated_secuence
                else:
                    sheetname = sheetname[:28] + deduplicated_secuence
            # Refeed the method until we get an unduplicated name
            return self._check_sheetname(sheetname, is_chartsheet=is_chartsheet)

    # "Short string"

    xlsxwriter.Workbook = PatchedXlsxWorkbook

except ImportError:
    _logger.debug("Can not import xlsxwriter`.")


class ReportXlsxAbstract(models.AbstractModel):
    _name = "report.report_xlsx.abstract"
    _description = "Abstract XLSX Report"

    def _get_objs_for_report(self, docids, data):
        """
        Returns objects for xlx report.  From WebUI these
        are either as docids taken from context.active_ids or
        in the case of wizard are in data.  Manual calls may rely
        on regular context, setting docids, or setting data.

        :param docids: list of integers, typically provided by
            qwebactionmanager for regular Models.
        :param data: dictionary of data, if present typically provided
            by qwebactionmanager for TransientModels.
        :param ids: list of integers, provided by overrides.
        :return: recordset of active model for ids.
        """
        if docids:
            ids = docids
        elif data and "context" in data:
            ids = data["context"].get("active_ids", [])
        else:
            ids = self.env.context.get("active_ids", [])
        return self.env[self.env.context.get("active_model")].browse(ids)

    def create_xlsx_report(self, docids, data):
        template_xml_id = self._name.replace('report.', '')
        xlsx_template = self.env.ref(template_xml_id, False)
        objs = self._get_objs_for_report(docids, data)
        file_data = BytesIO()

        if xlsx_template:  # write data to xlsx template
            decoded_data = base64.b64decode(xlsx_template.datas)
            workbook = load_workbook(BytesIO(decoded_data))
            self.generate_xlsx_report(workbook, data, objs)
            workbook.save(file_data)

        else:  # use xlsx writer to create xlsx file
            print('ok')
            workbook = xlsxwriter.Workbook(file_data, self.get_workbook_options())
            self.generate_xlsx_report(workbook, data, objs)
            workbook.close()

        file_data.seek(0)

        return file_data.read(), "xlsx"

    def preview_ms(self, docids, data):
        file_data_read = self.create_xlsx_report(docids, data)[0]
        attachment_data = base64.encodebytes(file_data_read)
        attachment = self.env['ir.attachment'].create({'name': self._name,
                                                       'datas': attachment_data,
                                                       'res_model': 'ir.actions.report',
                                                       'public': True})
        office_url = 'https://view.officeapps.live.com/op/view.aspx?src='
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true&filename=%s" \
              % (attachment.id, "Report")
        ms_url = office_url + quote((base_url + url), safe='~()*!.\'')
        preview_ms = self.env.ref('report_xlsx.action_report_ms_preview')
        # use cron_trigger in odoo 15
        cron_clean_attachment = self.env.ref('report_xlsx.clean_report_attachments_cron')
        cron_clean_attachment.sudo().nextcall = fields.datetime.now() + relativedelta(minutes=2)
        return preview_ms.report_action(attachment, data={'ms_url': ms_url})

    def get_workbook_options(self):
        """
        See https://xlsxwriter.readthedocs.io/workbook.html constructor options
        :return: A dictionary of options
        """
        return {}

    def generate_xlsx_report(self, workbook, data, objs):
        raise NotImplementedError()
