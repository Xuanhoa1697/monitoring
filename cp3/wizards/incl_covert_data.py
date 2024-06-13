# -*- coding: utf-8 -*-
from odoo import api, models, fields, _
import base64
import openpyxl
from io import BytesIO
import csv
import random
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from odoo.exceptions import ValidationError

class MonitoringInclConvert(models.TransientModel):
    _name = 'monitoring.incl.convert.data.wizard'

    file_data = fields.Many2many('ir.attachment', string="INCL Binary")
    date_from = fields.Date(string="Date", default=fields.Date.today)
    date_to = fields.Date(string="Date", default=fields.Date.today)
    random_number = fields.Boolean(string="Auto Number")
    for_date = fields.Date(string="Date", default=fields.Date.today)

    @api.model
    def default_get(self, fields_list):
        incl_id = self.env['inclinometter.file'].sudo().search([('attachment_create_ids', '!=', False)], limit=1)
        res = super(MonitoringInclConvert, self).default_get(fields_list)
        res['file_data'] = incl_id.attachment_create_ids.ids
        return res

    @api.onchange('date_from', 'date_to')
    def _onchange_date(self):
        if self.date_to < self.date_from and not self.random_number:
            raise ValidationError("Date to error!")
        if self.random_number:
            self.date_to = self.date_from

    @api.onchange('random_number')
    def _onchange_random_number(self):
        if self.random_number:
            self.date_to = self.date_from

    def convert_data(self):
        self.ensure_one()
        if not self.file_data:
            raise ValidationError('Vui lòng tải File Excel lên trước.')
        attachment_ids = []
        sheet_data = 'datas'
        start = self.date_from
        end = self.date_to if not self.random_number else start
        log = ''
        while start <= end:
            date = start.strftime('%d-%m-%Y')
            date_field = start.strftime('%m/%d/%Y') if not self.random_number else (self.for_date).strftime('%m/%d/%Y')
            datetime_field = self.random_time_in_office_hours()
            for attchment in self.file_data:
                # Ví dụ dữ liệu binary (base64 encoded) cho minh họa
                datas = []
                excel_binary_base64 = attchment.datas
                excel_binary = base64.b64decode(excel_binary_base64)
                file = BytesIO(excel_binary)
                workbook = openpyxl.load_workbook(file, read_only=True)
                sheet_names = workbook.sheetnames
                if '-' in date:
                    date = date.split("-")
                    if int(date[1]) < 10 and int(date[0]) > 10:
                        date[1] = date[1].replace("0",'')
                        date = "-".join(date)
                        if date not in sheet_names:
                            date = start.strftime('%d-%m-%Y')

                    elif int(date[1]) > 10 and int(date[0]) < 10:
                        date[0] = date[0].replace("0",'')
                        if date not in sheet_names:
                            date = start.strftime('%d-%m-%Y')
                    
                    elif int(date[1]) < 10 and int(date[0]) < 10:
                        is_check_date = False
                        if not is_check_date:
                            date[0] = date[0].replace("0",'')
                            if "-".join(date) in sheet_names:
                                date = "-".join(date)
                                is_check_date = True
                            else:
                                date = start.strftime('%d-%m-%Y')

                        if not is_check_date:
                            date = date.split("-")
                            date[1] = date[1].replace("0",'')
                            if "-".join(date) in sheet_names:
                                date = "-".join(date)
                                is_check_date = True
                            else:
                                date = start.strftime('%d-%m-%Y')

                        if not is_check_date:
                            date = date.split("-")
                            date[0] = date[0].replace("0",'')
                            date[1] = date[1].replace("0",'')
                            if "-".join(date) in sheet_names:
                                date = "-".join(date)
                                is_check_date = True
                            else:
                                date = start.strftime('%d-%m-%Y')
                    elif int(date[1]) < 10 and int(date[0]) == 10:
                        date[1] = date[1].replace("0",'')
                        if "-".join(date) in sheet_names:
                            date = "-".join(date)
                        else:
                            date = start.strftime('%d-%m-%Y')
                    
                    elif int(date[1]) == 10 and int(date[0]) < 10:
                        date[0] = date[0].replace("0",'')
                        if "-".join(date) in sheet_names:
                            date = "-".join(date)
                        else:
                            date = start.strftime('%d-%m-%Y')
                    else:
                        date = "-".join(date)

                if sheet_data not in sheet_names:
                    log += f'{str(date)} khong co sheet datas\n'
                    continue
                
                sheet_data_info = workbook[sheet_data]
                name = ''
                project = ''
                filename = ''
                probe = ''
                for row_data in sheet_data_info.iter_rows(values_only=True):
                    if row_data[0] == 'name':
                        name = row_data[1]
                    if row_data[0] == 'project':
                        project = row_data[1]
                    if row_data[0] == 'filename':
                        filename = row_data[1]
                    if row_data[0] == 'probe':
                        probe = row_data[1]

                if date in sheet_names:
                    sheet = workbook[date]
                    
                    for row in sheet.iter_rows(values_only=True):
                        if row[0] and '=L' in row[0]:
                            datas.insert(0, row)
                    
                    format_date_str = date_field.split("/")
                    format_date_str[2] = format_date_str[2][-2:]
                    format_date_substr = format_date_str[0]
                    format_date_str[0] = format_date_str[0][-1:] if int(format_date_str[0]) < 10 else format_date_str[0]
                    format_date_str[1] = format_date_str[1][-1:] if int(format_date_str[1]) < 10 else format_date_str[1]
                    data = [
                        ["***"],
                        ['GK 604M(v1.3.0.10',f'{format_date_substr + "/" + format_date_str[2]});2.0;FORMAT II'],
                        [f"PROJECT  :{project}"],
                        [f"HOLE NO. :{name}"],
                        [f"DATE     :{'/'.join(format_date_str)}"],
                        [f"TIME     :{str(datetime_field)}"],
                        [f"PROBE NO.:{probe}"],
                        [f"FILE NAME:{filename}"],
                        ["#READINGS:57"],
                        ["FLEVEL","    A+","    A-","    B+","    B+"]
                    ]
                    # if not 
                    current_space = 6
                    for item in datas:
                        item_11 = abs(float(item[11]))
                        len_item_index = len(str(item_11))
                        left_len_item_index = current_space - len_item_index
                        current_index_str = ''
                        for i in range(0, left_len_item_index):
                            current_index_str += " "
                        result_item_index = current_index_str + str(item_11)

                        item_13 = item[13] if not self.random_number else (item[13] + random.randint(-10, 10))
                        len_item_a0 = len(str(item_13))
                        left_len_item_a0 = current_space - len_item_a0
                        current_a0_str = ''
                        for i in range(0, left_len_item_a0):
                            current_a0_str += " "
                        result_item_a0 = current_a0_str + str(item_13)

                        item_14 = item[14] if not self.random_number else (item[14] + random.randint(-10, 10))
                        len_item_a180 = len(str(item_14))
                        left_len_item_a180 = current_space - len_item_a180
                        current_a180_str = ''
                        for i in range(0, left_len_item_a180):
                            current_a180_str += " "
                        result_item_a180 = current_a180_str + str(item_14)

                        item_15 = item[15] if not self.random_number else (item[15] + random.randint(-10, 10))
                        len_item_b0 = len(str(item_15))
                        left_len_item_b0 = current_space - len_item_b0
                        current_b0_str = ''
                        for i in range(0, left_len_item_b0):
                            current_b0_str += " "
                        result_item_b0 = current_b0_str + str(item_15)

                        item_16 = item[16] if not self.random_number else (item[16] + random.randint(-10, 10))
                        len_item_b180 = len(str(item_16))
                        left_len_item_b180 = current_space - len_item_b180
                        current_b180_str = ''
                        for i in range(0, left_len_item_b180):
                            current_b180_str += " "
                        result_item_b180 = current_b180_str + str(item_16)

                        data.append([result_item_index,result_item_a0,result_item_a180,result_item_b0,result_item_b180])
                    
                    with open(filename, 'w', newline='', encoding='utf-8') as csv_file:
                        # Khởi tạo writer object
                        csv_writer = csv.writer(csv_file)
                        
                        # Ghi dữ liệu vào file CSV
                        for row in data:
                            csv_writer.writerow(row)
                    with open(filename, 'rb') as csv_file:
                        csv_binary = csv_file.read()
                        csv_base64 = base64.b64encode(csv_binary)
                        
                        attachment_vals = {
                            'name': f'{str(date)}==' + filename,  # Tên của file đính kèm
                            'datas': csv_base64,  # Dữ liệu của file đính kèm (dưới dạng base64)
                            'res_model': self._name,  # Model mà file đính kèm được liên kết với
                            'res_id': self.id,  # ID của record mà file đính kèm được liên kết với
                        }
                        attachment = self.env['ir.attachment'].create(attachment_vals)
                        attachment_ids.append(attachment.id)
                if date not in sheet_names:
                    log += f'Không tìm thấy dữ liệu {str(date)} của file {name}\n'
            start += timedelta(days=1)
        if not attachment_ids:
            raise ValidationError(f'Hiện tại đang không phát sinh dữ liệu từ ngày {str(self.date_from)} đến ngày {str(self.date_to)}')  
        incl_id = self.env['inclinometter.file'].create({
            'name': f'INCL/{str(self.date_from)}',
            'attachment_ids': attachment_ids,
            'log': log,
            'attachment_create_ids': self.file_data
        })
        return {
            'name': _('Travel requested'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'inclinometter.file',
            'target': 'current',
            'res_id': incl_id.id
        }

    def random_time_in_office_hours(self):
        # Thời điểm bắt đầu và kết thúc giờ hành chính
        start_time = datetime.strptime('09:00:00', '%H:%M:%S')
        end_time = datetime.strptime('17:00:00', '%H:%M:%S')
        
        # Tính tổng số giây giữa thời điểm bắt đầu và kết thúc
        total_seconds = int((end_time - start_time).total_seconds())
        
        # Tạo một số giây ngẫu nhiên trong khoảng này
        random_seconds = random.randint(0, total_seconds)
        
        # Thêm số giây ngẫu nhiên vào thời điểm bắt đầu
        random_time = start_time + timedelta(seconds=random_seconds)
        
        return random_time.time()